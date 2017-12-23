import openpyxl, os
wb = openpyxl.load_workbook('myexcel.xlsx')
print(os.getcwd())
sheet = wb.get_sheet_by_name('工作表1')
print(type(sheet))
print(sheet.title)
print(sheet['A1'])
print(sheet['A1'].value)
print(type(sheet['A1'].value))
print(sheet.max_column, sheet.max_row)
print(sheet.cell(row=1, column=2).value)

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObject in rowOfCellObjects:
        print(cellObject.coordinate, cellObject.value)
    print('--- End of row ---')
